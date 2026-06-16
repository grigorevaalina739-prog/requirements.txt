import asyncio
import os
from datetime import date, datetime, timedelta
from io import BytesIO

import gspread
import pandas as pd
from google.oauth2.service_account import Credentials
from aiogram import Bot, Dispatcher, F, Router
from aiogram.filters import Command
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from apscheduler.schedulers.asyncio import AsyncIOScheduler
import aiosmtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BOT_TOKEN      = os.getenv("BOT_TOKEN")
DIRECTOR_ID    = int(os.getenv("DIRECTOR_ID", "5413167650"))
EMAIL_TO       = os.getenv("EMAIL_TO", "a.grigoryeva@miniso.kz")
SMTP_HOST      = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT      = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER      = os.getenv("SMTP_USER")
SMTP_PASSWORD  = os.getenv("SMTP_PASSWORD")
GOOGLE_CREDS   = os.getenv("GOOGLE_CREDENTIALS_JSON")

SHEETS = [
    ("1vAUE0ti9DM1DDD_uqnQ2mDOROavTtjKFccV9aOF_roE", "1008263264", "Задачи Борд NYYX"),
    ("1IgsFVUE42cT9YSRzZ27jYCkkCvKHyyijTLqnmFauF4E", "207788354",  "Стратегическая сессия 21-22 января 2026"),
    ("1eRz9DbQIj9PKPRTNnDR3m4p2AQQfKJLBJnJORxgJ0qs", "1708477262", "Протокол встречи по рекламе NYYX"),
    ("1TF56KEIgepd6u5khMuBt_mayktuQCEUX5eamaNToic4", "200182291",  "Протокол Встречи по ассортименту"),
    ("1gpwJqGpTl-Ttv9ieCOoW2e0w2J4qfgTXWeNfw7K998Q", "1397700586", "Встреча по SC NYYX"),
]

DONE_KEYWORDS = ["выполнено", "готово", "done", "завершено", "закрыто", "completed"]

EMPLOYEES = {
    "Кусайынова А":  0,
    "Мырзагали Е":   0,
    "Харужевский Е": 0,
    "Ахметова Т":    0,
    "Савченко С":    0,
    "Басембаева Д":  0,
    "AlinaNyyX":     5413167650,
}

COL_ALIASES = {
    "assignee": ["ответственный", "исполнитель", "ответственное лицо", "сотрудник"],
    "task":     ["задача", "название", "наименование"],
    "deadline": ["дедлайн", "срок", "срок выполнения", "срок исполнения", "дата"],
    "status":   ["статус", "выполнение", "состояние"],
    "priority": ["приоритет"],
    "comment":  ["комментарий", "примечание", "комментарии ответственных лиц"],
}

tasks_cache = []
pending_updates = {}

def get_gc():
    import json
    creds_dict = json.loads(GOOGLE_CREDS)
    scopes = ["https://spreadsheets.google.com/feeds",
              "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    return gspread.authorize(creds)

def get_sheet_by_gid(ss, gid):
    return next((s for s in ss.sheets() if str(s.id) == gid), None)

def find_col(headers, aliases):
    for alias in aliases:
        for i, h in enumerate(headers):
            if alias in h.lower():
                return i
    return -1

def fetch_all_tasks():
    gc = get_gc()
    all_tasks = []
    for sheet_id, gid, sheet_name in SHEETS:
        try:
            ss = gc.open_by_key(sheet_id)
            ws = get_sheet_by_gid(ss, gid)
            if not ws:
                continue
            rows = ws.get_all_values()
            if len(rows) < 2:
                continue
            headers = rows[0]
            col = {k: find_col(headers, v) for k, v in COL_ALIASES.items()}
            for i, row in enumerate(rows[1:], start=2):
                title = row[col["task"]] if col["task"] != -1 and col["task"] < len(row) else ""
                if not title.strip():
                    continue
                deadline_raw = row[col["deadline"]] if col["deadline"] != -1 and col["deadline"] < len(row) else ""
                deadline = parse_date(deadline_raw)
                status = row[col["status"]] if col["status"] != -1 and col["status"] < len(row) else ""
                all_tasks.append({
                    "sheet_id":   sheet_id,
                    "sheet_name": sheet_name,
                    "gid":        gid,
                    "row_num":    i,
                    "status_col": col["status"],
                    "assignee":   row[col["assignee"]].strip() if col["assignee"] != -1 and col["assignee"] < len(row) else "—",
                    "task":       title.strip(),
                    "deadline":   deadline,
                    "deadline_str": format_date(deadline),
                    "status":     status.strip(),
                    "priority":   row[col["priority"]].strip() if col["priority"] != -1 and col["priority"] < len(row) else "—",
                    "comment":    row[col["comment"]].strip() if col["comment"] != -1 and col["comment"] < len(row) else "",
                })
        except Exception as e:
            print(f"Ошибка {sheet_name}: {e}")
    return all_tasks

def update_task_status(sheet_id, gid, row_num, status_col, new_status):
    gc = get_gc()
    ss = gc.open_by_key(sheet_id)
    ws = get_sheet_by_gid(ss, gid)
    if ws and status_col != -1:
        cell = gspread.utils.rowcol_to_a1(row_num, status_col + 1)
        ws.update(cell, [[new_status]])

def parse_date(val):
    if not val:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            pass
    return None

def format_date(d):
    return d.strftime("%d.%m.%Y") if d else "—"

def is_done(status):
    return any(kw in status.lower() for kw in DONE_KEYWORDS)

def is_overdue(deadline):
    return bool(deadline and deadline < date.today())

def overdue_days(deadline):
    return (date.today() - deadline).days if deadline else 0

async def refresh_cache():
    global tasks_cache
    loop = asyncio.get_event_loop()
    tasks_cache = await loop.run_in_executor(None, fetch_all_tasks)
    print(f"Кэш обновлён: {len(tasks_cache)} задач")

class UpdateStatus(StatesGroup):
    waiting = State()

router = Router()
bot = Bot(token=BOT_TOKEN)

@router.message(Command("start"))
async def cmd_start(message: Message):
    is_dir = message.from_user.id == DIRECTOR_ID
    await message.answer(
        f"👋 Привет! Я агент контроля задач NYYX.\n"
        f"Роль: {'👑 Директор' if is_dir else '👤 Сотрудник'}\n\n"
        f"📌 Команды:\n"
        f"/tasks — мои задачи\n"
        f"/late — просроченные\n"
        f"/today — задачи на сегодня\n"
        f"/week — задачи на неделю\n"
        f"/summary — общая сводка\n"
        f"/report — Excel-отчёт\n\n"
        f"🔍 Напишите фамилию сотрудника для поиска"
    )

@router.message(Command("tasks"))
async def cmd_tasks(message: Message):
    uid = message.from_user.id
    assignee = next((n for n, tid in EMPLOYEES.items() if tid == uid), None)
    if not assignee and uid != DIRECTOR_ID:
        await message.answer("❌ Вы не найдены в списке сотрудников.")
        return
    tasks = [t for t in tasks_cache if not is_done(t["status"])]
    if assignee:
        tasks = [t for t in tasks if assignee.lower() in t["assignee"].lower()]
    if not tasks:
        await message.answer("✅ Нет активных задач!")
        return
    overdue = [t for t in tasks if is_overdue(t["deadline"])]
    active  = [t for t in tasks if not is_overdue(t["deadline"])]
    text = "📋 *Ваши задачи*\n\n"
    if overdue:
        text += f"🔴 *Просроченные ({len(overdue)}):*\n"
        for t in overdue:
            text += f"• {t['task'][:60]}\n  ⏰ {overdue_days(t['deadline'])} дн. | {t['sheet_name']}\n"
        text += "\n"
    if active:
        text += f"🔵 *Активные ({len(active)}):*\n"
        for t in active:
            text += f"• {t['task'][:60]}\n  📅 {t['deadline_str']} | {t['sheet_name']}\n"
    await message.answer(text, parse_mode="Markdown")

@router.message(Command("late"))
async def cmd_late(message: Message):
    if message.from_user.id != DIRECTOR_ID:
        await message.answer("❌ Только для директора.")
        return
    overdue = [t for t in tasks_cache if is_overdue(t["deadline"]) and not is_done(t["status"])]
    if not overdue:
        await message.answer("✅ Просроченных задач нет!")
        return
    by_assignee = {}
    for t in overdue:
        by_assignee.setdefault(t["assignee"], []).append(t)
    text = f"⚠️ *Просроченные задачи ({len(overdue)})*\n\n"
    for assignee, tasks in by_assignee.items():
        text += f"👤 *{assignee}* ({len(tasks)})\n"
        for t in tasks:
            text += f"  • {t['task'][:55]} — {overdue_days(t['deadline'])} дн.\n"
        text += "\n"
    await message.answer(text, parse_mode="Markdown")

@router.message(Command("today"))
async def cmd_today(message: Message):
    tasks = [t for t in tasks_cache if t["deadline"] == date.today() and not is_done(t["status"])]
    if not tasks:
        await message.answer(f"📅 На сегодня ({format_date(date.today())}) задач нет.")
        return
    text = f"📅 *Задачи на сегодня ({format_date(date.today())})*\n\n"
    for t in tasks:
        text += f"• *{t['assignee']}* — {t['task'][:55]}\n  {t['sheet_name']}\n"
    await message.answer(text, parse_mode="Markdown")

@router.message(Command("week"))
async def cmd_week(message: Message):
    end = date.today() + timedelta(days=7)
    tasks = [t for t in tasks_cache if t["deadline"] and date.today() <= t["deadline"] <= end and not is_done(t["status"])]
    if not tasks:
        await message.answer("📆 На неделю задач нет.")
        return
    text = f"📆 *Задачи на неделю ({len(tasks)})*\n\n"
    for t in sorted(tasks, key=lambda x: x["deadline"]):
        text += f"• {t['deadline_str']} | *{t['assignee']}* — {t['task'][:50]}\n"
    await message.answer(text, parse_mode="Markdown")

@router.message(Command("summary"))
async def cmd_summary(message: Message):
    if message.from_user.id != DIRECTOR_ID:
        await message.answer("❌ Только для директора.")
        return
    pending = [t for t in tasks_cache if not is_done(t["status"])]
    overdue = [t for t in pending if is_overdue(t["deadline"])]
    done    = [t for t in tasks_cache if is_done(t["status"])]
    by_sheet = {}
    for t in pending:
        by_sheet.setdefault(t["sheet_name"], []).append(t)
    text = (f"📊 *Сводка NYYX*\n\n"
            f"📌 Всего: {len(tasks_cache)}\n"
            f"✅ Выполнено: {len(done)}\n"
            f"🔵 В работе: {len(pending) - len(overdue)}\n"
            f"🔴 Просрочено: {len(overdue)}\n\n"
            f"*По проектам:*\n")
    for name, tasks in by_sheet.items():
        ov = sum(1 for t in tasks if is_overdue(t["deadline"]))
        text += f"📁 {name}: {len(tasks)}"
        if ov:
            text += f" ⚠️ {ov}"
        text += "\n"
    await message.answer(text, parse_mode="Markdown")

@router.message(Command("report"))
async def cmd_report(message: Message):
    await message.answer("⏳ Формирую отчёт...")
    pending = [t for t in tasks_cache if not is_done(t["status"])]
    excel   = build_excel(pending)
    today   = format_date(date.today()).replace(".", "_")
    from aiogram.types import BufferedInputFile
    await message.answer_document(
        BufferedInputFile(excel, filename=f"tasks_{today}.xlsx"),
        caption=f"📊 Невыполненные задачи\nВсего: {len(pending)} | Просрочено: {sum(1 for t in pending if is_overdue(t['deadline']))}"
    )

@router.message(F.text & ~F.text.startswith("/"))
async def search_handler(message: Message, state: FSMContext):
    current = await state.get_state()
    if current == UpdateStatus.waiting:
        await process_status_update(message, state)
        return
    query = message.text.strip().lower()
    if len(query) < 3:
        return
    results = [t for t in tasks_cache if query in t["assignee"].lower() or query in t["task"].lower()]
    pending = [t for t in results if not is_done(t["status"])]
    if not pending:
        await message.answer(f"🔍 По запросу *{message.text}* ничего не найдено.", parse_mode="Markdown")
        return
    by_assignee = {}
    for t in pending:
        by_assignee.setdefault(t["assignee"], []).append(t)
    for assignee, tasks in by_assignee.items():
        overdue = [t for t in tasks if is_overdue(t["deadline"])]
        active  = [t for t in tasks if not is_overdue(t["deadline"])]
        text = f"👤 *{assignee}*\n\n"
        if overdue:
            text += f"🔴 *Просроченные:*\n"
            for t in overdue:
                text += f"• {t['task'][:60]}\n  просрочка {overdue_days(t['deadline'])} дн. | {t['sheet_name']}\n"
            text += "\n"
        if active:
            text += f"🔵 *Активные:*\n"
            for t in active:
                text += f"• {t['task'][:60]}\n  📅 {t['deadline_str']} | {t['sheet_name']}\n"
        uid = message.from_user.id
        emp = next((n for n, tid in EMPLOYEES.items() if tid == uid), None)
        if emp and emp.lower() in assignee.lower():
            kb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text=f"✏️ {t['task'][:35]}", callback_data=f"upd_{i}")]
                for i, t in enumerate(tasks[:5])
            ])
            await message.answer(text, parse_mode="Markdown", reply_markup=kb)
        else:
            await message.answer(text, parse_mode="Markdown")

@router.callback_query(F.data.startswith("upd_"))
async def callback_update(call: CallbackQuery, state: FSMContext):
    idx = int(call.data.split("_")[1])
    uid = call.from_user.id
    emp = next((n for n, tid in EMPLOYEES.items() if tid == uid), None)
    if not emp:
        await call.answer("❌ Нет доступа")
        return
    user_tasks = [t for t in tasks_cache if emp.lower() in t["assignee"].lower() and not is_done(t["status"])]
    if idx >= len(user_tasks):
        await call.answer("Задача не найдена")
        return
    pending_updates[uid] = user_tasks[idx]
    await state.set_state(UpdateStatus.waiting)
    await call.message.answer(
        f"📝 Задача: *{user_tasks[idx]['task']}*\n\nНапишите новый статус:\n• В работе\n• Выполнено\n• На проверке",
        parse_mode="Markdown"
    )
    await call.answer()

async def process_status_update(message: Message, state: FSMContext):
    uid  = message.from_user.id
    task = pending_updates.get(uid)
    if not task:
        await state.clear()
        return
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, update_task_status,
        task["sheet_id"], task["gid"], task["row_num"], task["status_col"], message.text.strip())
    await message.answer(f"✅ Статус обновлён!\n*{task['task'][:60]}*\n→ {message.text.strip()}", parse_mode="Markdown")
    await state.clear()
    del pending_updates[uid]
    await refresh_cache()

def build_excel(tasks):
    rows = [{"Проект": t["sheet_name"], "Ответственный": t["assignee"], "Задача": t["task"],
             "Дедлайн": t["deadline_str"], "Статус": t["status"], "Приоритет": t["priority"],
             "Комментарий": t["comment"], "Просрочка": "Да" if is_overdue(t["deadline"]) else "Нет",
             "Дней просрочки": overdue_days(t["deadline"]) if is_overdue(t["deadline"]) else ""}
            for t in tasks]
    df  = pd.DataFrame(rows)
    buf = BytesIO()
    df.to_excel(buf, index=False, sheet_name="Задачи")
    buf.seek(0)
    wb  = load_workbook(buf)
    ws  = wb["Задачи"]
    hf  = PatternFill("solid", fgColor="1A237E")
    hfont = Font(color="FFFFFF", bold=True, name="Calibri")
    thin  = Side(style="thin", color="CCCCCC")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(horizontal="center"); cell.border = brd
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = brd; cell.font = Font(name="Calibri", size=10)
            if ws.cell(row=cell.row, column=8).value == "Да":
                cell.fill = PatternFill("solid", fgColor="FFF5F5")
    for i, w in enumerate([30,20,50,12,15,12,35,10,13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    out = BytesIO(); wb.save(out); return out.getvalue()

async def notify_employees():
    for name, tid in EMPLOYEES.items():
        if tid == 0:
            continue
        my_tasks = [t for t in tasks_cache if name.lower() in t["assignee"].lower() and not is_done(t["status"])]
        if not my_tasks:
            continue
        overdue = [t for t in my_tasks if is_overdue(t["deadline"])]
        active  = [t for t in my_tasks if not is_overdue(t["deadline"])]
        text = f"🌅 *Доброе утро!*\nВаши задачи на {format_date(date.today())}:\n\n"
        if overdue:
            text += f"🔴 *Просроченные ({len(overdue)}):*\n"
            for t in overdue:
                text += f"• {t['task'][:60]} — {overdue_days(t['deadline'])} дн.\n"
            text += "\n"
        if active:
            text += f"🔵 *Активные ({len(active)}):*\n"
            for t in active[:5]:
                text += f"• {t['task'][:60]} — {t['deadline_str']}\n"
        try:
            await bot.send_message(tid, text, parse_mode="Markdown")
        except Exception as e:
            print(f"Не удалось отправить {name}: {e}")

async def send_email_report():
    pending = [t for t in tasks_cache if not is_done(t["status"])]
    overdue = [t for t in pending if is_overdue(t["deadline"])]
    done    = [t for t in tasks_cache if is_done(t["status"])]
    today   = format_date(date.today())
    by_sheet = {}
    for t in pending:
        by_sheet.setdefault(t["sheet_name"], {}).setdefault(t["assignee"], []).append(t)
    html = f"""<div style="font-family:Arial,sans-serif;max-width:950px;margin:0 auto;">
    <h2 style="color:#1a237e;border-bottom:3px solid #1a237e;padding-bottom:8px;">
      📋 Невыполненные задачи — {today}</h2>
    <table style="border-collapse:collapse;margin-bottom:24px;"><tr>
      <td style="padding:12px 20px;background:#f3f4f6;border-radius:8px;text-align:center;">
        <div style="font-size:28px;font-weight:bold;color:#1a237e;">{len(pending)}</div>
        <div style="font-size:13px;color:#555;">Невыполненных</div></td>
      <td style="width:12px;"></td>
      <td style="padding:12px 20px;background:#fef2f2;border-radius:8px;text-align:center;">
        <div style="font-size:28px;font-weight:bold;color:#b91c1c;">{len(overdue)}</div>
        <div style="font-size:13px;color:#555;">Просрочено</div></td>
      <td style="width:12px;"></td>
      <td style="padding:12px 20px;background:#f0fdf4;border-radius:8px;text-align:center;">
        <div style="font-size:28px;font-weight:bold;color:#15803d;">{len(done)}</div>
        <div style="font-size:13px;color:#555;">Выполнено</div></td>
    </tr></table>"""
    for sheet_name, assignees in by_sheet.items():
        total = sum(len(v) for v in assignees.values())
        ov    = sum(1 for tasks in assignees.values() for t in tasks if is_overdue(t["deadline"]))
        html += f'<div style="margin-bottom:32px;"><h3 style="background:#1a237e;color:white;padding:10px 16px;border-radius:6px;">📁 {sheet_name} <span style="font-size:13px;font-weight:normal;">({total} задач{f" · ⚠️ {ov} просрочено" if ov else ""})</span></h3>'
        for assignee, tasks in assignees.items():
            aov = sum(1 for t in tasks if is_overdue(t["deadline"]))
            html += f'<div style="margin:10px 0 0 8px;"><div style="background:{"#fff3cd" if aov else "#e8f4fd"};border-left:4px solid {"#f59e0b" if aov else "#3b82f6"};padding:8px 14px;font-weight:bold;">👤 {assignee} <span style="font-weight:normal;font-size:12px;color:#666;">{len(tasks)} задач{f" · ⚠️ {aov}" if aov else ""}</span></div>'
            html += '<table style="border-collapse:collapse;width:100%;font-size:13px;"><tr style="background:#f8f9fa;"><th style="padding:6px 10px;text-align:left;border-bottom:1px solid #ddd;width:45%;">Задача</th><th style="padding:6px 10px;border-bottom:1px solid #ddd;width:12%;">Дедлайн</th><th style="padding:6px 10px;border-bottom:1px solid #ddd;">Статус</th><th style="padding:6px 10px;border-bottom:1px solid #ddd;">Комментарий</th></tr>'
            for i, t in enumerate(tasks):
                ov_t = is_overdue(t["deadline"])
                bg   = "#fff5f5" if ov_t else ("#ffffff" if i%2==0 else "#fafafa")
                html += f'<tr style="background:{bg};"><td style="padding:7px 10px;border-bottom:1px solid #eee;">{t["task"]}</td><td style="padding:7px 10px;border-bottom:1px solid #eee;color:{"#b91c1c" if ov_t else "#333"};font-weight:{"bold" if ov_t else "normal"};">{t["deadline_str"]}{" ⚠️" if ov_t else ""}</td><td style="padding:7px 10px;border-bottom:1px solid #eee;">{t["status"]}</td><td style="padding:7px 10px;border-bottom:1px solid #eee;color:#666;">{t["comment"]}</td></tr>'
            html += "</table></div>"
        html += "</div>"
    html += f"<p style='color:#ccc;font-size:11px;'>Отчёт сформирован автоматически {today}</p></div>"
    excel = build_excel(pending)
    msg = MIMEMultipart("mixed")
    msg["From"] = SMTP_USER; msg["To"] = EMAIL_TO
    msg["Subject"] = f"📋 Невыполненные задачи — {today}"
    msg.attach(MIMEText(html, "html", "utf-8"))
    att = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    att.set_payload(excel); encoders.encode_base64(att)
    att.add_header("Content-Disposition", "attachment", filename=f"tasks_{today.replace('.','_')}.xlsx")
    msg.attach(att)
    try:
        await aiosmtplib.send(msg, hostname=SMTP_HOST, port=SMTP_PORT,
                              username=SMTP_USER, password=SMTP_PASSWORD, start_tls=True)
        print(f"Email отправлен на {EMAIL_TO}")
    except Exception as e:
        print(f"Ошибка email: {e}")

async def main():
    dp = Dispatcher(storage=MemoryStorage())
    dp.include_router(router)
    scheduler = AsyncIOScheduler(timezone="Asia/Almaty")
    scheduler.add_job(refresh_cache,     "interval", minutes=30, id="cache")
    scheduler.add_job(notify_employees,  "cron", hour=9,  minute=0,  id="notify")
    scheduler.add_job(send_email_report, "cron", hour=9,  minute=15, id="email")
    scheduler.start()
    await refresh_cache()
    print("Агент запущен!")
    await dp.start_polling(bot, allowed_updates=["message", "callback_query"])

if __name__ == "__main__":
    asyncio.run(main())
